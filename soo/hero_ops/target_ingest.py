"""히어로 목표·준비물량 → 스타일(base 품번)별 정규화.

소스: ★MSTRD 상품기획 xlsx (Drive의 Office 파일). Sheets API로는 못 읽어
      Drive get_media 로 내려받아 openpyxl(read_only)로 읽는다. (PLM xlsx 읽기와 같은 방식)

'히어로목표' 탭 = 전치(품번이 열) + 채널 2블록:
  - R2="ON 목표"  → C~Y열  (온라인 목표 블록)
  - R2="OFF 목표" → AB~AX열 (오프라인 목표 블록)  ※ 같은 품번이 양 블록에 등장. TOTAL = ON+OFF.
  - 각 블록 R3=품번, R4=품명, R5=준비물량, R6=목표소진율, R7=목표판매량, R8~=월별.
  - 그 뒤(BC~)에 목표 거래액(GMV) 블록도 있음 — qty 섹션은 R3=="GMV" 만나면 종료. (거래액은 추후)

반환: { base_style_no: {
          target_qty, target_qty_on, target_qty_off,
          prep_qty,  prep_qty_on,  prep_qty_off,
          target_sellthrough } }
  (목표 미설정 스타일은 dict에 없음 → 소비측에서 '목표 미설정' 처리)
"""
from __future__ import annotations

import io
import re

TARGET_FILE_ID = "1VTf5psWNm0-EuR1gz4AQrOTy7r5j0ebS"   # ★MSTRD_26SS 상품기획(상품MAP).xlsx
TARGET_TAB = "히어로목표"
STYLE_RE = re.compile(r"^M[A-Z0-9]{8}$")

# rows[] 0-indexed (min_row=1부터): R3=품번, R5=준비물량, R6=목표소진율, R7=목표판매량
_R_STYLE, _R_PREP, _R_SELL, _R_TGT = 2, 4, 5, 6


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _base(style) -> str:
    return str(style).strip().split("-")[0]


def load_workbook(drive):
    """Drive에서 xlsx 내려받아 openpyxl 워크북 반환 (read_only, 값만)."""
    data = drive.files().get_media(fileId=TARGET_FILE_ID, supportsAllDrives=True).execute()
    import openpyxl
    return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)


def parse_targets(drive) -> dict:
    wb = load_workbook(drive)
    ws = wb[TARGET_TAB]
    rows = list(ws.iter_rows(min_row=1, max_row=8, max_col=80, values_only=True))
    r2, r3 = rows[1], rows[_R_STYLE]
    r5, r6, r7 = rows[_R_PREP], rows[_R_SELL], rows[_R_TGT]
    wb.close()

    on: dict[str, dict] = {}
    off: dict[str, dict] = {}
    block = None
    for c in range(2, len(r3)):
        lab = str(r2[c]).strip() if c < len(r2) and r2[c] else ""
        cell = str(r3[c]).strip() if c < len(r3) and r3[c] else ""
        if lab == "ON 목표":
            block = on
        elif lab == "OFF 목표":
            block = off
        if cell == "GMV" or cell.endswith("거래액"):   # qty 섹션 끝 (그 뒤는 목표 거래액)
            break
        if block is None or not STYLE_RE.match(cell):
            continue
        block[_base(cell)] = {
            "prep": _num(r5[c]) if c < len(r5) else None,
            "target": _num(r7[c]) if c < len(r7) else None,
            "sell": _num(r6[c]) if c < len(r6) else None,
        }

    out: dict[str, dict] = {}
    for base in set(on) | set(off):
        o, f = on.get(base, {}), off.get(base, {})
        tq_on, tq_off = o.get("target"), f.get("target")
        pq_on, pq_off = o.get("prep"), f.get("prep")
        out[base] = {
            "target_qty_on": tq_on,
            "target_qty_off": tq_off,
            "target_qty": (tq_on or 0) + (tq_off or 0) or None,
            "prep_qty_on": pq_on,
            "prep_qty_off": pq_off,
            "prep_qty": (pq_on or 0) + (pq_off or 0) or None,
            "target_sellthrough": o.get("sell") or f.get("sell"),
        }
    return out


if __name__ == "__main__":   # 단독 실행 = 파싱 + 검증
    import sys
    from pathlib import Path
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
    from soo.auth import get_credentials, build_services
    ROOT = Path(__file__).resolve().parents[2]
    drive = build_services(get_credentials(ROOT / "credentials.json", ROOT / "token.json"))["drive"]

    t = parse_targets(drive)
    print(f"목표 보유 품번 {len(t)}개")
    print(f"{'품번':12} {'목표(ON)':>8} {'목표(OFF)':>8} {'목표TOTAL':>9} {'준비TOTAL':>9} {'소진율':>6}")
    for k, v in sorted(t.items()):
        print(f"{k:12} {(v['target_qty_on'] or 0):8.0f} {(v['target_qty_off'] or 0):8.0f} "
              f"{(v['target_qty'] or 0):9.0f} {(v['prep_qty'] or 0):9.0f} {(v['target_sellthrough'] or 0):6.3f}")
    # 검산: 목표 ≈ 준비 × 소진율
    bad = [k for k, v in t.items() if v["target_qty"] and v["prep_qty"]
           and abs(v["target_qty"] - v["prep_qty"] * (v["target_sellthrough"] or 0)) > max(5, v["target_qty"] * 0.03)]
    print(f"\n검산(목표≈준비×소진율) 불일치: {len(bad)}개 {bad[:5]}")
