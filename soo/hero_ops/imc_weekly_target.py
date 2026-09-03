# -*- coding: utf-8 -*-
"""IMC 운영계획 시트2의 주차 그리드(I~AH) 자동 주입 — **목표 판매량**과 **신규 입고**.

두 블록을 서로 다른 원천에서 채운다.
  ① 목표 판매량 (R14/19/24/29/34/39/44/49/54) ← `26FW HERO 일자별 목표 셋팅`.xlsx (아래 상세)
  ② 신규 입고   (R12/17/22/27/32/37/42/47/52) ← `무탠본부_오더시트` `MD투입` 탭
       헤더 R7 기준 **최종품번(B) · 타겟시즌(AI) · 현/실 입고일(AT) · 예상 입고량(AU)**.
       **타겟시즌 = 2026FW 인 행만** 집계한다(같은 스타일이 다른 시즌으로도 발주돼 있음).
       대상 STY는 26FW 히어로 실적 대시보드의 **히어로별 탭 B열**(R13~)에서 동적으로 읽는다
       — 탭 이름이 곧 시트2 C열 히어로명이라 이름이 바뀌면 verify_layout 이 잡아준다.

──────────────────────────────────────────────────────────────────────────────
① 목표 판매량 원천: `26FW HERO 일자별 목표 셋팅` (박은진님, **.xlsx**) — Drive에 Office 파일로 있다.
      xlsx라 IMPORTRANGE 대상이 안 되므로(구글시트 네이티브만 지원) 매 실행마다
      get_media 로 내려받아 openpyxl 로 캐시값을 읽는다. 즉 **원본 파일을 고치면 다음 실행에 반영**된다.
      (원본을 '새 파일로 업로드'하면 파일 ID가 바뀌어 옛 파일을 계속 읽게 되니 그때만 SRC_FILE_ID 교체.)

타깃: `무신사스탠다드 히어로 IMC 운영 계획 - 26FW` 시트2 — 주차 열 I~AH(마감일 7/5~12/27, 26주).
      히어로 블록의 **수량 행(목표 판매량·신규 입고)만** 쓴다.

★ 금액 칸은 절대 건드리지 않는다 (2026-07-29 결정).
  시트2에는 금액 칸이 두 종류 섞여 있다.
    - 히어로 블록의 `목표 매출액` 행(R15/20/…)
    - 오프라인·온라인 `주차별 목표` 블록(R57~68 / R105~115) — 서식 `#,###,,` = **원 단위 저장·백만원 표시**.
      (수량을 넣으면 화면상 0으로 보여 값이 안 들어간 것처럼 됨. 실제로 그 사고를 한 번 겪었다.)
  이 자동화는 수량(pcs)만 다루므로 위 영역은 읽지도 쓰지도 않는다.

★ 원천 두 탭이 일부 어긋난다 (2026-07-29 확인):
    `주차별 목표 셋팅`(→ `목표 그래프`) = 산출 원본. 자연PLC → 매출비 → 조정 주간매출 → 목표 판매량.
    `일자별 목표 셋팅`          = 그 주차 목표를 일자로 배분한 파생물.
  커브드 팬츠 **오프라인**만 7월 1~4주에서 일자별이 주차별보다 636~709 적고, 7월5주 이후로는
  매주 정확히 +84 많다(합계 -762 = 0.5%). 나머지 5개 시리즈는 온·오프 전 주차 일치.
  → 주차 그리드에 넣을 값이므로 **정본인 주차별을 기본(--source weekly)** 으로 쓴다.

★ 주차 정렬 함정: 원천 주차 *라벨*("7월2주")과 시트2 주차 열은 한 칸씩 어긋난다
  (원천 7월1주 = 7/1~7/5 = 시트2 I열, 원천 7월2주 = 7/6~7/12 = 시트2 J열).
  라벨 순서로 붙이면 전 구간이 밀리므로 **주 마감일(끝 날짜)로 매칭**한다.
  시트2의 한 주 = [마감일-6, 마감일]. 원천 12월5주(12/28~12/31)는 시트2 범위 밖이라 버린다.

★ 커버리지(목표): 원천 시리즈는 6종(커브드 팬츠/빅토리아 울/라이트다운/그리드·메시 플리스/에센셜 플리스/웜 팬츠).
  시트2의 **힛탠다드·리커버리는 원천에 없어** 건드리지 않는다(빈칸 유지). TOTAL 행도 6종 합계다.
  또 원천은 시리즈별 **핵심 스타일만** 대상이다(예: 커브드 6스타일·준비 328,731 vs 시트2 신규입고 723,900)
  — 목표수량과 준비물량의 모수가 다르므로 소진율을 이 둘로 계산하면 안 된다.

──────────────────────────────────────────────────────────────────────────────
② 신규 입고 주의사항 (2026-07-30 확인)
  - 6/29(첫 주 시작) **이전 입고분은 주차 범위 밖**이라 버린다. 커브드팬츠가 115,679장으로 대부분인데,
    26FW 타겟시즌인데도 6월에 이미 입고된 물량이다. 하반기 그리드에 억지로 넣지 않는다.
  - AT(입고일)가 비어 있는 행도 버린다(라이트다운 2,000장). 날짜를 모르면 주차를 정할 수 없다.
  - **리커버리는 2026FW 행 30건이 있으나 AU(예상 입고량)가 전부 0** — 수량 미입력 상태다.
    합계가 0인 히어로는 0을 쓰지 않고 통째로 건너뛴다(빈칸 유지). 나중에 수량이 채워지면 자동 반영된다.
  - 시트2 F열(신규입고 누적)은 RAW1 준비물량이라 여기 합계와 모수가 다르다(TOTAL 3,003,605 vs 2,041,276).
    캐리오버·타 시즌 발주가 F열에는 들어 있기 때문 — 두 값이 안 맞는 게 정상이다.

안전 규칙(0으로 덮어쓰는 사고 방지):
  1. 원천 파싱 결과가 비었거나 전부 0이면 **아무것도 쓰지 않고 중단**한다.
  2. 시트2의 행 라벨(C열 히어로명·D열 지표명)을 매번 검증해 레이아웃이 바뀌었으면 중단한다.
  3. 원천에 없는 히어로 행은 스킵(빈칸 유지) — 0을 채워 '목표 0'처럼 보이게 하지 않는다.
  4. 주차별 탭의 시리즈 집합이 일자별 탭과 다르면 중단(한쪽에만 추가된 시리즈를 놓치지 않기 위함).
"""
from __future__ import annotations

import argparse
import datetime as dt
import io
import re
import sys
import traceback
from pathlib import Path

# ── 원천/타깃 ────────────────────────────────────────────────────────────────
SRC_FILE_ID = "1CB10ouLsOZplJuPoSOkkAXhvzwgtR0zD"     # 26FW HERO 일자별 목표 셋팅 (.xlsx)
SRC_TAB = "일자별 목표 셋팅"
SRC_WEEK_TAB = "목표 그래프"     # 주차별 목표(정본)의 시리즈×주차×ON/OFF 정리본

IMC_SHEET_ID = "1jDRvZncF0D2RoeCGdxNso3wrUz09BQPk4a7JEhs1ElQ"
IMC_TAB = "시트2"

# ── 신규 입고 원천 ───────────────────────────────────────────────────────────
ORDER_SHEET_ID = "13R4gcJ7cDlReY-vwjXZf0kMZ7tC4kr2-S7PC9uziVUQ"   # 무탠본부_오더시트
ORDER_TAB = "MD투입"
ORDER_HEADER_ROW = 7            # 데이터는 R8부터
ORDER_SEASON = "2026FW"         # AI(타겟시즌) 필터 — 같은 STY가 타 시즌으로도 발주돼 있다
# 헤더 텍스트(공백·개행 제거 후 비교) → 쓰임. 열 위치는 매번 헤더에서 찾는다.
# ★2026-09-03: "최종품번"은 STY 가 아니라 **옵션코드(품번-컬러)** 다. 8/19 원천이 바뀌면서
#   `MMFPC3A15` → `MMFPC3D81-BK` 꼴이 되어 대시보드 STY 와 교집합 0 → 입고가 매일 통째로 실패했다.
#   → ①STY 열은 **후보 라벨을 순서대로** 찾고(품번 우선, 최종품번 폴백) ②값은 `-` 앞을 잘라
#     항상 STY 로 정규화한다(sales_rollup._base 와 같은 규칙). 원천이 어느 쪽으로 돌아가도 맞는다.
ORDER_COLS = {"sty": ("품번", "최종품번", "대표품번"), "season": "타겟시즌",
              "date": "현/실입고일", "qty": "예상입고량",
              "real_date": "실입고일", "real_qty": "실입고량"}

# 히어로별 대상 STY = 26FW 히어로 실적 대시보드의 히어로 탭 B열(R13~)
DASH_SHEET_ID = "1-A04_TwKZJNPkFg27USkKAScZRu6CAhbgVeXk9c09nA"
DASH_STY_RANGE = "B13:B"
# 기입고물량 = 히어로 탭 합계행(R12)의 **HU(1/1 예측 재고)** + **6/1~그리드 직전까지의 입고**.
#   ★ 처음엔 HU + HX(입고량, 물류입고기준)로 잡았는데, HX는 '오늘까지' 누계라 그리드 시작 시점이
#     아니라 현재 보유량이어서 지난 주차와 겹쳤다. 입고분은 오더시트에서 기간을 잘라 쓴다
#     (사용자 지시 2026-07-30: 입고는 6/1~6/28). 이러면 그리드(6/29~)와 구간이 안 겹친다.
DASH_TOTAL_ROW = 12
DASH_PREIN_HEADER_ROW = 8
DASH_PREIN_COLS = ("1/1예측재고",)   # 공백·개행 제거 후 비교
DASH_PREIN_SCAN = "HP{r}:IF{r}"      # 헤더/합계행에서 위 컬럼을 찾을 범위
PREIN_FROM = dt.date(2026, 6, 1)     # 기입고물량에 포함할 입고 시작일 (끝은 그리드 첫 주 시작 전날)

SEASON_YEAR = 2026

# ── 원천 레이아웃 (일자별 목표 셋팅) ─────────────────────────────────────────
# ★ 아래는 **폴백 기본값일 뿐**, 실제 위치는 `_locate_src()` 가 라벨로 찾는다.
#   담당자가 앞쪽에 보조 컬럼을 끼워 넣으면 통째로 밀린다 — 2026-08-04 실제 발생:
#   F~I에 '채널별 일자별 매출목표 얼라인' 4열 삽입 → 주차 F→J · 라벨/날짜 G→K · 데이터 H→L.
#   그 결과 라벨열 K가 데이터 열로 잡혀 시리즈 'HERO 시리즈'가 생겼고 시리즈 집합 가드에 걸려 중단됐다.
#   (같은 파일을 읽는 `target_26fw.py` 는 같은 날 `_locate()` 로 고쳤는데 여기는 빠져 있었다.)
SRC_COL0 = 8            # H열부터 스타일×채널 블록
SRC_R_SERIES, SRC_R_LINE, SRC_R_CH, SRC_R_STYLE = 5, 6, 7, 8
SRC_DAILY_ROW0 = 18     # 일자 그리드 시작
SRC_C_WEEK, SRC_C_DATE = 6, 7          # F열 = 주차 라벨, G열 = 날짜("07월 01일" 텍스트)

# 라벨열 판별용 — 이 라벨이 한 열에 모두 있으면 그 열이 라벨열(=날짜열), 데이터는 그 다음 열부터.
SRC_LABEL_KEYS = ("HERO시리즈", "라인", "채널", "품번")
SRC_ROW_LABELS = {"SRC_R_SERIES": "HERO시리즈", "SRC_R_LINE": "라인",
                  "SRC_R_CH": "채널", "SRC_R_STYLE": "품번"}
_WEEK_RE = re.compile(r"\d{1,2}\s*월\s*\d\s*주")

# ── 타깃 레이아웃 (시트2) ────────────────────────────────────────────────────
# ★ 주차 열 위치는 하드코딩하지 않는다. 2026-07-30에 I열에 '기입고물량'이 삽입되면서
#   주차 그리드가 I~AH → J~AI 로 통째로 밀린 적이 있다(기존 값도 같이 밀려 정합은 유지됐다).
#   매 실행마다 마감일 행에서 날짜 시리얼이 연속된 구간을 찾아 주차 열로 쓴다.
WEEK_ROW = 7                    # 마감일 행
# 주차 그리드 왼쪽의 기초(그리드 이전) 물량 열. 라벨은 **부분 일치**로 찾는다 —
# 실제로 '기입고물량' → '시점재고+기입고물량(7월1주전까지)' 로 바뀐 적이 있다.
PREIN_LABEL = "기입고물량"
_SERIAL_MIN, _SERIAL_MAX = 45000, 48000     # 2023~2031 — 날짜 시리얼 판별용

# 히어로 블록: (C열 히어로 라벨, 신규입고 행, 누적입고재고 행, 목표판매량 행).
# 블록 구성 = 신규입고 / 누적입고재고 / 목표판매량 / 목표매출액 / 누적소진율 (5행).
# 금액 행(목표 매출액)과 누적 소진율은 의도적으로 제외 — 수량 세 행만 쓴다.
BLOCKS = [
    ("TOTAL", 12, 13, 14),
    ("커브드팬츠", 17, 18, 19),
    ("라이트다운", 22, 23, 24),
    ("힛탠다드", 27, 28, 29),
    ("빅토리아 울", 32, 33, 34),
    ("그리드/메시 플리스", 37, 38, 39),
    ("에센셜 플리스", 42, 43, 44),
    ("웜 팬츠", 47, 48, 49),
    ("리커버리", 52, 53, 54),
]
HERO_LABELS = [b[0] for b in BLOCKS if b[0] != "TOTAL"]   # 대시보드 탭 이름과 동일

# 표기 흔들림 흡수: 공백 제거 후 비교 + 별칭
ALIAS = {"샤기/플러피플리스": "에센셜플리스"}
_EPOCH = dt.date(1899, 12, 30)          # 구글시트 날짜 시리얼 기준
_DATE_RE = re.compile(r"(\d{1,2})\D+(\d{1,2})")


def norm(s) -> str:
    """히어로/시리즈명 정규화 — 공백 전부 제거 후 별칭 치환."""
    k = re.sub(r"\s+", "", str(s or ""))
    return ALIAS.get(k, k)


def _num(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _to_date(v):
    """원천 G열 → date. '07월 01일' 텍스트와 datetime 둘 다 받는다."""
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    m = _DATE_RE.search(str(v or ""))
    if not m:
        return None
    try:
        return dt.date(SEASON_YEAR, int(m.group(1)), int(m.group(2)))
    except ValueError:
        return None


def _locate_src(ws) -> dict:
    """일자별 탭의 라벨열·헤더행·주차열 위치를 **시트에서 직접 찾는다**(하드코딩 오프셋 금지).

    라벨열 = 위쪽 30행 안에 'HERO 시리즈'·'라인'·'채널'·'품번'이 모두 있는 열.
             이 열은 데이터 행에서 날짜("07월 01일")를 담는 열이기도 하다. 데이터는 그 다음 열부터.
    주차열 = 라벨열 왼쪽에서 "7월1주" 꼴이 5개 이상 나오는 첫 열.
    못 찾으면 모듈 폴백 상수를 쓰되, 어느 항목이 폴백인지 로그에 남긴다.
    """
    scan_r = min(ws.max_row, 30)
    scan_c = min(ws.max_column, 40)
    loc, fell_back = {}, []

    c_label = None
    for c in range(1, scan_c + 1):
        labels = {re.sub(r"\s+", "", str(ws.cell(r, c).value or ""))
                  for r in range(1, scan_r + 1)}
        if all(k in labels for k in SRC_LABEL_KEYS):
            c_label = c
            break
    if c_label is None:
        c_label, _ = SRC_C_DATE, fell_back.append("라벨열")
    loc["c_date"] = c_label
    loc["c_data0"] = c_label + 1

    rowof = {}
    for r in range(1, scan_r + 1):
        v = re.sub(r"\s+", "", str(ws.cell(r, c_label).value or ""))
        if v and v not in rowof:
            rowof[v] = r
    for key, lab in SRC_ROW_LABELS.items():
        if lab in rowof:
            loc[key] = rowof[lab]
        else:
            loc[key], _ = globals()[key], fell_back.append(lab + "행")

    # 일별 그리드 시작 = 라벨열이 날짜로 읽히는 첫 행(헤더 아래)
    r0 = None
    for r in range(max(loc["SRC_R_STYLE"], SRC_R_STYLE) + 1, min(ws.max_row, 60) + 1):
        if _to_date(ws.cell(r, c_label).value):
            r0 = r
            break
    if r0 is None:
        r0, _ = SRC_DAILY_ROW0, fell_back.append("일자시작행")
    loc["r_daily0"] = r0

    # 주차열 = 라벨열 왼쪽에서 "7월1주" 꼴이 5개 이상인 첫 열
    probe = range(r0, min(ws.max_row, r0 + 60) + 1)
    c_week = None
    for c in range(c_label - 1, 0, -1):
        hits = sum(1 for r in probe if _WEEK_RE.fullmatch(str(ws.cell(r, c).value or "").strip()))
        if hits >= 5:
            c_week = c
            break
    if c_week is None:
        c_week, _ = SRC_C_WEEK, fell_back.append("주차열")
    loc["c_week"] = c_week
    loc["fell_back"] = fell_back
    return loc


def _col_letter(idx: int) -> str:
    """1-indexed 열 번호 → A1 표기."""
    s = ""
    while idx:
        idx, r = divmod(idx - 1, 26)
        s = chr(65 + r) + s
    return s


# ── 1) 원천 읽기 ─────────────────────────────────────────────────────────────
def load_source(drive):
    """xlsx를 내려받아 (일자별 열, 주차별 목표, 주차 구간) 파싱.

    반환:
      cols     : [{series, line, ch, style, daily:{date: qty}}]        (일자별 탭)
      weekly   : {(series_norm, ch): {주차라벨: qty}}                   (목표 그래프 탭)
      wk_range : {주차라벨: (시작일, 종료일)}                            (일자별 탭 F/G열)
    """
    import openpyxl
    from googleapiclient.http import MediaIoBaseDownload

    buf = io.BytesIO()
    dl = MediaIoBaseDownload(buf, drive.files().get_media(fileId=SRC_FILE_ID,
                                                          supportsAllDrives=True),
                             chunksize=4 * 1024 * 1024)
    done = False
    while not done:
        _, done = dl.next_chunk()
    buf.seek(0)

    wb = openpyxl.load_workbook(buf, data_only=True)
    for tab in (SRC_TAB, SRC_WEEK_TAB):
        if tab not in wb.sheetnames:
            raise RuntimeError(f"원천에 '{tab}' 탭이 없습니다 (탭: {wb.sheetnames})")

    # --- 일자별 탭 ---
    ws = wb[SRC_TAB]
    loc = _locate_src(ws)
    print(f"[원천 레이아웃] 주차 {_col_letter(loc['c_week'])}열 · 날짜/라벨 {_col_letter(loc['c_date'])}열 "
          f"· 데이터 {_col_letter(loc['c_data0'])}열~ · 일자 시작 R{loc['r_daily0']}"
          + (f"  ⚠ 폴백: {', '.join(loc['fell_back'])}" if loc["fell_back"] else ""))

    cols, blank_series = [], []
    for c in range(loc["c_data0"], ws.max_column + 1):
        series, ch = ws.cell(loc["SRC_R_SERIES"], c).value, ws.cell(loc["SRC_R_CH"], c).value
        style = str(ws.cell(loc["SRC_R_STYLE"], c).value or "").strip()
        if not series or not ch:
            # 시리즈 라벨만 빠진 열은 조용히 흘리지 않고 경고한다
            # (2026-08-04 실제: AL열 라이트다운 키즈 online MKEDJ9K01 의 시리즈 칸이 비어 있었다).
            if ch and style:
                blank_series.append(f"{_col_letter(c)}({style} {str(ch).strip()})")
            continue
        cols.append({"c": c, "series": str(series).strip(),
                     "line": str(ws.cell(loc["SRC_R_LINE"], c).value or "").strip(),
                     "ch": str(ch).strip().lower(),
                     "style": style,
                     "daily": {}})
    if blank_series:
        print(f"  ⚠ 시리즈 라벨이 비어 제외한 열 {len(blank_series)}개: {', '.join(blank_series)}"
              f" — 원천 확인 필요(기본 소스가 weekly라 목표값 자체엔 영향 없음)")

    wk_range: dict[str, tuple[dt.date, dt.date]] = {}
    for r in range(loc["r_daily0"], ws.max_row + 1):
        d = _to_date(ws.cell(r, loc["c_date"]).value)
        if d is None:
            continue
        wk = str(ws.cell(r, loc["c_week"]).value or "").strip()
        if wk:
            a, b = wk_range.get(wk, (d, d))
            wk_range[wk] = (min(a, d), max(b, d))
        for col in cols:
            q = _num(ws.cell(r, col["c"]).value)
            if q:
                col["daily"][d] = col["daily"].get(d, 0.0) + q

    # --- 목표 그래프 탭: [시리즈명] / [주차|ONLINE|OFFLINE|TOTAL] / 주차행... 블록 반복 ---
    g = wb[SRC_WEEK_TAB]
    weekly: dict[tuple[str, str], dict[str, float]] = {}
    for r in range(1, g.max_row + 1):
        name = g.cell(r, 1).value
        if not name or str(g.cell(r + 1, 1).value or "").strip() != "주차":
            continue
        key = norm(name)
        rr = r + 2
        while True:
            wk = str(g.cell(rr, 1).value or "").strip()
            if not wk.endswith("주"):
                break
            weekly.setdefault((key, "online"), {})[wk] = _num(g.cell(rr, 2).value)
            weekly.setdefault((key, "offline"), {})[wk] = _num(g.cell(rr, 3).value)
            rr += 1

    # 두 탭의 시리즈 집합이 어긋나면 **그 시리즈만 빼고** 나머지는 주입한다.
    #   ★2026-08-13 변경: 종전엔 raise 로 통째 중단이라, 담당자가 '일자별 목표 셋팅'에 힛탠다드를
    #     넣고 정본 '목표 그래프' 블록을 아직 안 만든 이틀(8/12·8/13) 동안 **목표도 입고도 전혀
    #     주입되지 않았다**. 한쪽에만 있는 시리즈는 어차피 어느 기준으로 써도 반쪽이므로
    #     제외하는 게 맞지만, 그 때문에 멀쩡한 6종까지 멈출 이유는 없다.
    #   조용히 흘리지 않기 위해 notes 로 올려 로그·슬랙에 남긴다(호출부에서 처리).
    s_daily = {norm(c["series"]) for c in cols}
    s_week = {k[0] for k in weekly}
    notes = []
    if s_daily != s_week:
        only_d, only_w = sorted(s_daily - s_week), sorted(s_week - s_daily)
        drop = set(only_d) | set(only_w)
        if only_d:
            notes.append(f"'{SRC_TAB}'에만 있고 정본 '{SRC_WEEK_TAB}'에 블록이 없음: {', '.join(only_d)}")
        if only_w:
            notes.append(f"'{SRC_WEEK_TAB}'에만 있음: {', '.join(only_w)}")
        cols = [c for c in cols if norm(c["series"]) not in drop]
        weekly = {k: v for k, v in weekly.items() if k[0] not in drop}
    return cols, weekly, wk_range, notes


# ── 1-b) 신규 입고 원천 (무탠본부_오더시트 MD투입) ──────────────────────────
def load_hero_styles(sheets) -> dict[str, str]:
    """26FW 대시보드 히어로 탭 B열 → {STY: 히어로명}. 중복 STY는 먼저 나온 히어로에 귀속.

    ★죽은 블록은 세지 않는다(2026-08-01). 히어로가 다른 품목으로 옮겨가면 옛 탭에 **품번만 남은 빈 블록**이
    남는데(컬러키 C열이 비어 있고 헤더는 '0SKU'), 그걸 소유로 세는 바람에 실제 소유 탭과 충돌해
    "여러 히어로 탭에 같은 STY" 오류로 매일 주입이 멈췄다(MWFNPAA09 커브드팬츠↔웜 팬츠).
    → **컬러키(C열 `STY-XX`)가 하나라도 살아 있는 탭만** 그 STY의 소유로 본다.
    """
    res = sheets.spreadsheets().values().batchGet(
        spreadsheetId=DASH_SHEET_ID,
        ranges=[f"'{h}'!A13:C400" for h in HERO_LABELS]).execute()["valueRanges"]
    live = {}          # (hero, sty) → 살아있는 컬러키 수
    for hero, vr in zip(HERO_LABELS, res):
        for r in vr.get("values", []):
            sty = str(r[1]).strip() if len(r) > 1 else ""
            key = str(r[2]).strip() if len(r) > 2 else ""
            if sty and key.startswith(sty + "-"):
                live[(hero, sty)] = live.get((hero, sty), 0) + 1
    sty2hero, dupes = {}, []
    for hero, vr in zip(HERO_LABELS, res):
        for r in vr.get("values", []):
            sty = str(r[1]).strip() if len(r) > 1 else ""
            if not sty or not live.get((hero, sty)):
                continue                      # 빈(이동·폐기) 블록 — 소유로 세지 않는다
            if sty in sty2hero and sty2hero[sty] != hero:
                dupes.append(f"{sty}({sty2hero[sty]}↔{hero})")
            sty2hero.setdefault(sty, hero)
    if not sty2hero:
        raise RuntimeError("26FW 대시보드 히어로 탭에서 STY를 하나도 못 읽었습니다 — 주입 중단.")
    if dupes:
        raise RuntimeError("여러 히어로 탭에 같은 STY가 있습니다 — 귀속이 모호해 중단:\n  "
                           + ", ".join(dupes))
    return sty2hero


def load_prein(sheets) -> dict[str, float]:
    """히어로 탭 합계행의 HU(1/1 예측재고). 기입고물량의 '재고' 몫.

    입고 몫(6/1~그리드 직전)은 오더시트에서 따로 잘라 더한다 — load_inbound 의 stat['prein'].
    열 위치는 헤더행(R8)의 라벨로 찾는다 — 대시보드 컬럼이 밀려도 따라가기 위함.
    """
    ranges = []
    for h in HERO_LABELS:
        ranges.append(f"'{h}'!" + DASH_PREIN_SCAN.format(r=DASH_PREIN_HEADER_ROW))
        ranges.append(f"'{h}'!" + DASH_PREIN_SCAN.format(r=DASH_TOTAL_ROW))
    res = sheets.spreadsheets().values().batchGet(
        spreadsheetId=DASH_SHEET_ID, ranges=ranges,
        valueRenderOption="UNFORMATTED_VALUE").execute()["valueRanges"]

    out, missing, broken = {}, [], []
    for i, hero in enumerate(HERO_LABELS):
        hdr = (res[2 * i].get("values") or [[]])[0]
        tot = (res[2 * i + 1].get("values") or [[]])[0]
        labels = [re.sub(r"\s+", "", str(c)) for c in hdr]
        val = 0.0
        for want in DASH_PREIN_COLS:
            if want not in labels:
                missing.append(f"{hero}:{want}")
                continue
            j = labels.index(want)
            raw = tot[j] if j < len(tot) else 0
            # ★ 수식 오류(#REF! 등)를 0으로 삼키면 안 된다 — 대시보드 라이트다운 탭에서
            #   실제로 HX가 #REF!였고, 그때 _num()이 조용히 0을 돌려줬다.
            if isinstance(raw, str) and raw.strip().startswith("#"):
                broken.append(f"{hero}:{want}={raw.strip()[:24]}")
                continue
            val += _num(raw)
        out[hero] = val
    if missing:
        raise RuntimeError("대시보드 히어로 탭에서 기입고물량 컬럼을 못 찾았습니다 — 주입 중단:\n  "
                           + ", ".join(missing))
    if broken:
        raise RuntimeError("대시보드 기입고물량 셀이 수식 오류입니다 — 0으로 덮지 않도록 중단:\n  "
                           + ", ".join(broken))
    out["TOTAL"] = sum(out.values())
    return out


def _sty_key(v) -> str:
    """오더시트 품번 칸 → STY. 옵션코드(`품번-컬러`)면 `-` 앞을 취한다.

    ★히어로 STY 자체엔 `-` 가 없다(대시보드 B열 9자리). sales_rollup._base 와 같은 규칙.
    """
    return str(v).strip().split("-")[0]


def load_inbound(sheets, sty2hero, weeks, basis="hybrid", as_of=None, shift_overdue=False):
    """MD투입 타겟시즌 2026FW 행 → {히어로: 주차별 입고수량}. TOTAL 포함.

    basis='hybrid'(기본): **지난 주차는 실입고(AV/AW), 앞은 예상(AT/AU)**.
      AT/AU만 쓰면 이행되지 않은 과거 예정일이 입고된 것처럼 잡힌다
      (실례: 웜 팬츠 MKCNPAZ01·MKDNPAZ01 예정일 7/24이 지났으나 AW=0, 물류입고도 0).
      예정일이 지났는데 실입고가 없는 물량은 '지연'으로 집계해 로그로 알린다
      (shift_overdue=True면 현재 주차로 옮긴다 — 기본은 원래 예정 주차 유지).
    basis='plan': 종전처럼 전 구간 AT/AU(예상)만 사용.
    """
    hdr = sheets.spreadsheets().values().get(
        spreadsheetId=ORDER_SHEET_ID,
        range=f"'{ORDER_TAB}'!A{ORDER_HEADER_ROW}:CZ{ORDER_HEADER_ROW}").execute().get("values", [[]])
    labels = [re.sub(r"\s+", "", str(c)) for c in (hdr[0] if hdr else [])]
    idx = {}
    for key, name in ORDER_COLS.items():
        cands = (name,) if isinstance(name, str) else tuple(name)
        hit = next((c for c in cands if c in labels), None)
        if hit is None:
            raise RuntimeError(f"{ORDER_TAB} R{ORDER_HEADER_ROW}에서 '{'/'.join(cands)}' 열을 "
                               f"못 찾았습니다 — 오더시트 구조 변경. 주입 중단.")
        idx[key] = labels.index(hit)
        if key == "sty":
            print(f"  · STY 열 = '{hit}' ({_col_letter(idx[key] + 1)})")

    order = list(ORDER_COLS)
    res = sheets.spreadsheets().values().batchGet(
        spreadsheetId=ORDER_SHEET_ID,
        ranges=[f"'{ORDER_TAB}'!{_col_letter(idx[k] + 1)}{ORDER_HEADER_ROW + 1}:"
                f"{_col_letter(idx[k] + 1)}" for k in order],
        valueRenderOption="UNFORMATTED_VALUE").execute()["valueRanges"]
    col = {k: vr.get("values", []) for k, vr in zip(order, res)}

    def val(k, i):
        rows = col[k]
        r = rows[i] if i < len(rows) else []
        return r[0] if r else ""

    n = max(len(v) for v in col.values())
    per = {h: [0.0] * len(weeks) for h in HERO_LABELS}
    stat = {"rows": 0, "before": {}, "prein": {}, "after": {}, "nodate": {}, "total": {},
            "actual": {}, "plan": {}, "overdue": {},
            # 진단용 — 0건일 때 "시즌이 없나 / 키가 안 맞나"를 로그만 보고 가른다
            "season_rows": 0, "sample": []}
    w_start, w_end = weeks[0][0], weeks[-1][1]
    as_of = as_of or dt.date.today()
    now_wi = next((wi for wi, (_, b) in enumerate(weeks) if b >= as_of), None)

    for i in range(n):
        if str(val("season", i)).strip() != ORDER_SEASON:
            continue
        stat["season_rows"] += 1
        key = _sty_key(val("sty", i))
        hero = sty2hero.get(key)
        if hero is None:
            if key and len(stat["sample"]) < 8 and key not in stat["sample"]:
                stat["sample"].append(key)
            continue
        stat["rows"] += 1

        rq, rd = _num(val("real_qty", i)), _num(val("real_date", i))
        if basis == "hybrid" and rq > 0 and rd > 0:
            qty, d, kind = rq, _EPOCH + dt.timedelta(days=int(rd)), "actual"
        else:
            qty = _num(val("qty", i))
            if not qty:
                continue
            serial = _num(val("date", i))
            if serial <= 0:
                stat["nodate"][hero] = stat["nodate"].get(hero, 0.0) + qty
                continue
            d = _EPOCH + dt.timedelta(days=int(serial))
            # 예정일이 지났는데 실입고가 없다 = 지연. 물량을 버리지 않고 현재 주차로 모은다.
            kind = "overdue" if (basis == "hybrid" and d <= as_of) else "plan"

        stat["total"][hero] = stat["total"].get(hero, 0.0) + qty
        stat[kind][hero] = stat[kind].get(hero, 0.0) + qty

        # 지연(예정일 경과·실입고 미기입)을 현재 주차로 옮길지는 선택.
        #   ★ 오더시트 AV/AW는 히어로마다 관리 편차가 크다 — 빅토리아 울은 물류입고(HX) 112,677인데
        #     실입고 열이 통째로 비어 있다. 그래서 '옮기기'를 기본으로 두면 그 물량이 현재 주차에
        #     몰려 가짜 스파이크가 생긴다. 기본은 원래 예정 주차에 그대로 두고 로그로만 알린다.
        if kind == "overdue" and shift_overdue:
            if now_wi is None:              # 기준일이 그리드 끝을 지났다
                stat["after"][hero] = stat["after"].get(hero, 0.0) + qty
            else:
                per[hero][now_wi] += qty
            continue
        if d < w_start:
            # 그리드 직전 구간(PREIN_FROM~w_start-1)은 '기입고물량'의 입고 몫으로 넘긴다.
            # 그보다 이른 입고는 어디에도 넣지 않고 로그로만 알린다.
            key = "prein" if d >= PREIN_FROM else "before"
            stat[key][hero] = stat[key].get(hero, 0.0) + qty
            continue
        if d > w_end:                       # 그리드 이후(연말 밖) — 넣을 칸이 없다
            stat["after"][hero] = stat["after"].get(hero, 0.0) + qty
            continue
        for wi, (a, b) in enumerate(weeks):
            if a <= d <= b:
                per[hero][wi] += qty
                break

    out = {h: per[h] for h in HERO_LABELS if any(per[h])}
    total = [0.0] * len(weeks)
    for arr in out.values():
        for i, v in enumerate(arr):
            total[i] += v
    if out:
        out["TOTAL"] = total
    return out, stat


def cumulative(inbound: dict, prein: dict, weeks, mode: str, as_of: dt.date) -> dict:
    """누적 입고 재고.

    mode='plan'(기본) : 기입고물량 + 전 주차 누계.
                        기입고물량의 입고 몫이 그리드 직전(6/28)까지로 잘려 있어 구간이 겹치지 않는다.
    mode='forward'    : 기입고물량 + 오늘 이후 주차만 누계.
                        기입고물량을 '오늘까지 누계'로 잡던 시절의 이중계상 회피용 — 지금은 불필요.
    """
    out = {}
    for hero, arr in inbound.items():
        acc = prein.get(hero, 0.0)
        run = []
        for (_, end), v in zip(weeks, arr):
            if mode == "plan" or end > as_of:
                acc += v
            run.append(acc)
        out[hero] = run
    # 입고 스케줄은 없지만 기입고물량만 있는 히어로도 누적은 그려준다.
    for hero, v in prein.items():
        if hero not in out and v:
            out[hero] = [v] * len(weeks)
    return out


# ── 2) 주차 경계 (시트2 마감일) ──────────────────────────────────────────────
def load_grid(sheets) -> dict:
    """시트2 마감일 행을 훑어 주차 열 구간과 '기입고물량' 열을 찾는다.

    반환 {weeks: [(시작, 마감)], c0: 첫 주차 열(1-indexed), prein: 기입고물량 열 or None}.
    열을 고정하지 않는 이유는 상단 주석 참고(중간에 열이 삽입된 이력이 있다).
    """
    row = (sheets.spreadsheets().values().get(
        spreadsheetId=IMC_SHEET_ID, range=f"'{IMC_TAB}'!A{WEEK_ROW}:BZ{WEEK_ROW}",
        valueRenderOption="UNFORMATTED_VALUE").execute().get("values", [[]]) or [[]])[0]

    prein = next((i + 1 for i, v in enumerate(row)
                  if PREIN_LABEL in re.sub(r"\s+", "", str(v))), None)

    best: list[int] = []
    cur: list[int] = []
    for i, v in enumerate(row):
        n = _num(v)
        if _SERIAL_MIN <= n <= _SERIAL_MAX:
            cur.append(i + 1)
        else:
            if len(cur) > len(best):
                best = cur
            cur = []
    if len(cur) > len(best):
        best = cur
    if len(best) < 10:
        raise RuntimeError(f"시트2 {WEEK_ROW}행에서 마감일(날짜) 구간을 못 찾았습니다 — 주입 중단.")
    if best != list(range(best[0], best[0] + len(best))):
        raise RuntimeError(f"마감일 열이 연속이 아닙니다: {best} — 주입 중단.")

    weeks = []
    for c in best:
        end = _EPOCH + dt.timedelta(days=int(_num(row[c - 1])))
        weeks.append((end - dt.timedelta(days=6), end))
    return {"weeks": weeks, "c0": best[0], "prein": prein}


def map_weeks(weeks, wk_range) -> list[str | None]:
    """시트2 주차 → 원천 주차 라벨. 주 마감일(끝 날짜)이 같은 것을 붙인다."""
    by_end = {b: wk for wk, (a, b) in wk_range.items()}
    out = [by_end.get(end) for _, end in weeks]
    if not any(out):
        raise RuntimeError("시트2 주차와 원천 주차가 하나도 매칭되지 않습니다 — 주입 중단.")
    return out


# ── 3) 레이아웃 검증 ────────────────────────────────────────────────────────
def verify_layout(sheets) -> None:
    """행 라벨이 기대와 다르면 즉시 중단 — 엉뚱한 행에 숫자를 쓰는 사고 방지."""
    vals = sheets.spreadsheets().values().get(
        spreadsheetId=IMC_SHEET_ID, range=f"'{IMC_TAB}'!A1:D60").execute().get("values", [])

    def cell(r, c):
        row = vals[r - 1] if r - 1 < len(vals) else []
        return str(row[c - 1]).strip() if c - 1 < len(row) else ""

    bad = []
    for hero, inb_r, cum_r, qty_r in BLOCKS:
        if norm(cell(inb_r, 3)) != norm(hero):          # 블록 첫 행(C열)에 히어로명
            bad.append(f"R{inb_r} C열 히어로명 '{cell(inb_r, 3)}' ≠ '{hero}'")
        for row, want in ((inb_r, "신규 입고"), (cum_r, "누적 입고 재고"), (qty_r, "목표 판매량")):
            if cell(row, 4) != want:
                bad.append(f"R{row} D열 '{cell(row, 4)}' ≠ '{want}'")
    if bad:
        raise RuntimeError("시트2 레이아웃이 바뀌었습니다 — 주입 중단:\n  " + "\n  ".join(bad))


# ── 4) 집계 ─────────────────────────────────────────────────────────────────
def aggregate(cols, weekly, wk_map, weeks, mode):
    """{시리즈: 주차별 목표수량 리스트} — 온·오프 합산. 'TOTAL' 키에 전체 합."""
    n = len(weeks)
    series_all = sorted({norm(c["series"]) for c in cols})
    out: dict[str, list[float]] = {}

    for s in series_all:
        arr = [0.0] * n
        for ch in ("online", "offline"):
            for wi, (a, b) in enumerate(weeks):
                if mode == "weekly":
                    lbl = wk_map[wi]
                    arr[wi] += weekly.get((s, ch), {}).get(lbl, 0.0) if lbl else 0.0
                else:
                    arr[wi] += sum(v for col in cols
                                   if norm(col["series"]) == s and col["ch"] == ch
                                   for d, v in col["daily"].items() if a <= d <= b)
        if any(arr):
            out[s] = arr

    total = [0.0] * n
    for arr in out.values():
        for i, v in enumerate(arr):
            total[i] += v
    if any(total):
        out["TOTAL"] = total
    return out


# ── 5) 주입 ─────────────────────────────────────────────────────────────────
def build_updates(grid, qty: dict | None, inbound: dict | None,
                  cum: dict | None, opening: dict | None) -> tuple[list[dict], dict]:
    """수량 세 행(+기입고물량 열)만 배치 업데이트로 만든다.

    값이 없는 히어로는 건드리지 않는다(빈칸 유지 — 0을 써서 '입고 없음'처럼 보이게 하지 않는다).
    """
    c0, n = grid["c0"], len(grid["weeks"])
    rng0, rng1 = _col_letter(c0), _col_letter(c0 + n - 1)
    data = []
    skipped = {"목표 판매량": [], "신규 입고": [], "누적 입고 재고": []}
    for hero, inb_r, cum_r, qty_r in BLOCKS:
        for src, row, label in ((qty, qty_r, "목표 판매량"),
                                (inbound, inb_r, "신규 입고"),
                                (cum, cum_r, "누적 입고 재고")):
            if src is None:
                continue
            # 목표는 정규화 키(공백 제거), 입고는 히어로 라벨 그대로가 키다.
            arr = src.get(norm(hero), src.get(hero))
            if arr is None:
                skipped[label].append(hero)
                continue
            data.append({"range": f"'{IMC_TAB}'!{rng0}{row}:{rng1}{row}",
                         "values": [[round(v) for v in arr]]})
        # 기입고물량 열 — 신규 입고 행에만 쓴다(누적 행의 시작값으로도 이미 반영돼 있다).
        if opening is not None and grid["prein"] and hero in opening:
            pc = _col_letter(grid["prein"])
            data.append({"range": f"'{IMC_TAB}'!{pc}{inb_r}",
                         "values": [[round(opening[hero])]]})
    return data, skipped


def main() -> int:
    from soo.auth import build_services, get_credentials

    # 로컬 콘솔이 cp949라 ⚠ 같은 기호에서 죽는 것 방지 (CI는 UTF-8이라 무해).
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    ap = argparse.ArgumentParser(description="IMC 시트2 주차 그리드 주입 (목표 판매량 · 신규 입고)")
    ap.add_argument("--only", choices=["all", "target", "inbound"], default="all",
                    help="all=둘 다(기본) / target=목표 판매량만 / inbound=신규 입고만")
    ap.add_argument("--source", choices=["weekly", "daily"], default="weekly",
                    help="목표 소스: weekly=주차별 목표(정본, 기본) / daily=일자별 목표를 주간 합산")
    ap.add_argument("--basis", choices=["hybrid", "plan"], default="hybrid",
                    help="신규 입고 기준: hybrid=지난 주차는 실입고(AV/AW)+앞은 예상(AT/AU) (기본) / "
                         "plan=전 구간 예상(AT/AU)")
    ap.add_argument("--shift-overdue", action="store_true",
                    help="예정일이 지났는데 실입고 미기입인 물량을 현재 주차로 옮긴다 "
                         "(기본은 원래 예정 주차 유지 — 오더시트 실입고 열 관리 편차 때문)")
    ap.add_argument("--cum", choices=["plan", "forward"], default="plan",
                    help="누적 입고 재고: plan=기입고물량+전 주차 누계(기본) / "
                         "forward=기입고물량+오늘 이후 주차만")
    ap.add_argument("--as-of", default="", help="누적 기준일 (YYYY-MM-DD, 기본 오늘)")
    ap.add_argument("--dry-run", action="store_true", help="시트에 쓰지 않고 결과만 출력")
    args = ap.parse_args()

    root = Path(__file__).resolve().parents[2]
    svc = build_services(get_credentials(root / "credentials.json", root / "token.json"))
    sheets, drive = svc["sheets"], svc["drive"]

    verify_layout(sheets)
    grid = load_grid(sheets)
    weeks = grid["weeks"]
    print(f"주차 {len(weeks)}주 ({weeks[0][0]} ~ {weeks[-1][1]}) · "
          f"열 {_col_letter(grid['c0'])}~{_col_letter(grid['c0'] + len(weeks) - 1)} · "
          f"기입고물량 열 {_col_letter(grid['prein']) if grid['prein'] else '없음'}")

    qty = inbound = cum = opening = None
    warns: list[str] = []      # 주입은 됐지만 사람이 알아야 하는 것 (슬랙 통지)
    hard: list[str] = []       # 한 축이 통째로 실패 — 종료코드 1

    # ── ① 목표 판매량 ──
    #   ★원천 격리(2026-08-13): 목표가 죽어도 ② 입고는 주입한다. 종전엔 여기서 raise 하면
    #     입고까지 같이 멈췄다(실제로 8/12·8/13 이틀 전량 미주입).
    if args.only in ("all", "target"):
        try:
            cols, weekly, wk_range, notes = load_source(drive)
            warns += [f"목표 원천 시리즈 불일치 — 제외하고 주입: {n}" for n in notes]
            for n in notes:
                print(f"  ⚠ {n} → 해당 시리즈 제외하고 나머지만 주입")
            if not cols:
                raise RuntimeError("목표 원천에서 스타일×채널 열을 하나도 못 읽었습니다 — 목표 주입 중단.")
            if sum(sum(c["daily"].values()) for c in cols) <= 0:
                raise RuntimeError("목표 원천 수량 합계가 0입니다 — 목표 주입 중단(0으로 덮어쓰기 방지).")
            wk_map = map_weeks(weeks, wk_range)
            qty = aggregate(cols, weekly, wk_map, weeks, args.source)
            unmapped = [str(e) for (_, e), m in zip(weeks, wk_map) if m is None]
            dropped = sorted(set(wk_range) - {m for m in wk_map if m})
            print(f"\n[목표 판매량] 소스={args.source} · 원천 열 {len(cols)}개")
            if unmapped:
                print(f"  ⚠ 원천 주차를 못 찾은 시트2 주차: {', '.join(unmapped)}")
            if dropped:
                print(f"  · 시트2 범위 밖이라 버린 원천 주차: {', '.join(dropped)}")
        except Exception as e:
            qty = None
            hard.append(f"목표 판매량 주입 실패 — {type(e).__name__}: {e}")
            print(f"\n[목표 판매량] ★실패 — {type(e).__name__}: {e}")
            traceback.print_exc()
            print("  → 목표 행은 건드리지 않고 입고만 진행한다.")

    # ── ② 신규 입고 ── (①과 같은 이유로 격리)
    if args.only in ("all", "inbound"):
        try:
            sty2hero = load_hero_styles(sheets)
            as_of = dt.date.fromisoformat(args.as_of) if args.as_of else dt.date.today()
            inbound, stat = load_inbound(sheets, sty2hero, weeks, args.basis, as_of, args.shift_overdue)
            if not inbound:
                raise RuntimeError(
                    f"{ORDER_TAB}에서 {ORDER_SEASON} 입고를 하나도 못 읽었습니다 — 입고 주입 중단.\n"
                    f"  · {ORDER_SEASON} 행 {stat['season_rows']}건 · STY 매칭 0건 "
                    f"(대상 STY {len(sty2hero)}개)\n"
                    f"  · 원천 STY 예시: {', '.join(stat['sample']) or '(비어 있음)'}\n"
                    f"  · 대시보드 STY 예시: {', '.join(list(sty2hero)[:5])}\n"
                    f"  → 둘의 형식이 다르면 오더시트 품번 열이 바뀐 것이다.")
            carry = load_prein(sheets)                    # 1/1 예측재고 (대시보드 HU)
            opening = {}                                  # 기입고물량 = 재고 + 6/1~그리드직전 입고
            for h in HERO_LABELS:
                v = carry.get(h, 0.0) + stat["prein"].get(h, 0.0)
                if v:
                    opening[h] = v
            opening["TOTAL"] = sum(opening.values())
            cum = cumulative(inbound, opening, weeks, args.cum, as_of)
            print(f"\n[신규 입고] {ORDER_SEASON} · 기준={args.basis} · 대상 STY {len(sty2hero)}개 "
                  f"· 매칭 행 {stat['rows']}건")
            ov = "지연→현재주차" if args.shift_overdue else "지연(예정주차 유지)"
            for lbl, key in (("실입고", "actual"), ("예상", "plan"), (ov, "overdue")):
                d = stat.get(key) or {}
                if d:
                    print(f"  · {lbl} {sum(d.values()):,.0f} — "
                          + ", ".join(f"{h} {v:,.0f}" for h, v in sorted(d.items())))
            pre = {h: v for h, v in opening.items() if v and h != "TOTAL"}
            print(f"  · 기입고물량 = 1/1 예측재고 + 입고({PREIN_FROM}~{weeks[0][0] - dt.timedelta(days=1)})"
                  f" → {_col_letter(grid['prein']) if grid['prein'] else '열없음'}열, 합 {opening['TOTAL']:,.0f}")
            for h, v in sorted(pre.items()):
                print(f"      {h:<16} 재고 {carry.get(h, 0):>8,.0f} + 입고 {stat['prein'].get(h, 0):>9,.0f}"
                      f" = {v:>9,.0f}")
            if not grid["prein"] and pre:
                print(f"  ⚠ '{PREIN_LABEL}' 열이 없어 기입고물량을 못 씁니다 — 누적 행에만 반영됩니다.")
            print(f"  · 누적 방식 = {args.cum} (기준일 {as_of})")
            for label, d in ((f"{PREIN_FROM} 이전이라 어디에도 미반영", stat["before"]),
                             ("그리드 이후라 제외", stat["after"]),
                             ("입고일 없어 제외", stat["nodate"])):
                if d:
                    print(f"  · {label}: " + ", ".join(f"{h} {v:,.0f}" for h, v in sorted(d.items())))
            zero = [h for h in HERO_LABELS if not stat["total"].get(h)]
            if zero:
                print(f"  ⚠ 입고수량 0(미입력)이라 신규입고 행은 건드리지 않음: {', '.join(zero)}")
        except Exception as e:
            inbound = cum = opening = None
            hard.append(f"신규 입고 주입 실패 — {type(e).__name__}: {e}")
            print(f"\n[신규 입고] ★실패 — {type(e).__name__}: {e}")
            traceback.print_exc()
            print("  → 입고·누적 행은 건드리지 않고 목표만 진행한다.")

    data, skipped = build_updates(grid, qty, inbound, cum, opening)
    print(f"\n갱신 셀 {len(data)}건 (수량 행만 — 금액 칸 미변경)")
    for label, heroes in skipped.items():
        if heroes:
            print(f"  · {label} 미기입: {', '.join(heroes)}")
    for hero, inb_r, cum_r, qty_r in BLOCKS:
        t = (qty or {}).get(norm(hero))
        b = (inbound or {}).get(hero)
        c = (cum or {}).get(hero)
        if t is None and b is None:
            continue
        print(f"  {hero:<16} 목표 {sum(t) if t else 0:>9,.0f} (R{qty_r})"
              f" · 입고 {sum(b) if b else 0:>9,.0f} (R{inb_r})"
              f" · 기입고 {(opening or {}).get(hero, 0):>8,.0f}"
              f" · 누적끝 {c[-1] if c else 0:>9,.0f} (R{cum_r})")

    if args.dry_run:
        print("\n[dry-run] 시트에 쓰지 않았습니다.")
        _report(warns, hard, dry=True)
        return 1 if hard else 0

    if data:
        sheets.spreadsheets().values().batchUpdate(
            spreadsheetId=IMC_SHEET_ID,
            body={"valueInputOption": "RAW", "data": data}).execute()
        print(f"\n✅ 시트2 주입 완료 — {len(data)}개 행 × {len(weeks)}주 (금액 칸 미변경)")
    else:
        print("\n쓸 값이 없습니다 — 시트 미변경.")

    _report(warns, hard)
    # 한 축이 통째로 죽었으면 실패로 끝낸다(워크플로 슬랙 알림이 받는다).
    # 시리즈 일부 제외 같은 부분 문제는 주입은 됐으므로 0으로 끝내되 아래에서 슬랙을 따로 보낸다.
    return 1 if hard else 0


def _report(warns: list[str], hard: list[str], dry: bool = False) -> None:
    """주입 후 요약 — 경고는 슬랙 DM으로, 하드 실패는 로그만(워크플로 failure 훅이 알린다)."""
    for h in hard:
        print(f"\n★ {h}")
    for w in warns:
        print(f"\n⚠ {w}")
    if warns and not hard and not dry:
        # 부분 주입은 초록불로 끝나므로 여기서 직접 알리지 않으면 아무도 모른다.
        try:
            from soo.hero_ops import notify
            notify.send("⚠️ IMC 주차별 목표 — 원천 불일치로 일부 시리즈를 빼고 주입했습니다.\n"
                        + "\n".join(f"· {w}" for w in warns)
                        + f"\n\n원천 `26FW HERO 일자별 목표 셋팅` 정본 탭('{SRC_WEEK_TAB}')에 "
                          "해당 시리즈 블록이 생기면 자동으로 다시 잡힙니다.")
        except Exception as e:      # 알림 실패가 주입을 되돌리진 않는다
            print(f"슬랙 통지 실패: {type(e).__name__}: {e}")


if __name__ == "__main__":
    raise SystemExit(main())
