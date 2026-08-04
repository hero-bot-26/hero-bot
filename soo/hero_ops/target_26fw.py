"""26FW 히어로 목표(일자별) → 스타일(base 품번)별 기간 누적 목표 + 준비수량.

소스: Drive 파일 **`26FW HERO 일자별 목표 셋팅`**(.xlsx, 담당 eunjin.park) `일자별 목표 셋팅` 탭.
      기존 26SS 목표 소스(`히어로목표(거래량)` 탭)와 **별도** — 그 탭은 1/1~ 달력 기준이라
      26FW 누계(7/1~)와 비교하면 무조건 미달로 보였다(사용자 지적 2026-07-30).

`일자별 목표 셋팅` 레이아웃 (G열=라벨, H열~=품번×채널 컬럼):
  R2 판매개시행(자동) · R3 판매개시일 · R4 목표소진율 · R5 HERO 시리즈 · R6 라인(남성/여성/키즈)
  R7 채널(online/offline) · R8 품번 · R9 품명 · R10 준비수량 · R11~R16 월별(7~12월)
  R17 시즌목표판매량 · **R18~ 일별 목표**(F열=주차 라벨, G열=날짜 '07월 01일', H열~=일별 수량)

기간 윈도우 = 매출 쪽과 동일 규약(데이터는 전일까지):
  YTD(=26FW 시즌 누계 FWTD) = [시즌 시작(=시트 첫 일자), as_of-1] · MTD = [월초, as_of-1]
  WEEK = [as_of-7, as_of-1] · DAY = [as_of-1]

반환: { base 품번: {tq: {기간: {t,o,f}}, prep: {t,o,f}, sellthrough} }  ← sales_rollup 목표 구조와 동일.
★목표가 설정된 히어로만 들어온다(현재 S/A급 6종: 커브드팬츠·라이트다운·빅토리아울·웜팬츠·
  그리드/메시 플리스·에센셜 플리스). 나머지는 dict에 없어 화면에서 '목표 미설정'.
"""
from __future__ import annotations

import datetime
import io
import re

# Drive 파일 ID (소스 레지스트리 키 `target_26fw` 로 교체 가능)
TARGET_FID = "1CB10ouLsOZplJuPoSOkkAXhvzwgtR0zD"
TARGET_TAB = "일자별 목표 셋팅"

PERIODS = ["YTD", "MTD", "WEEK", "DAY"]
STYLE_RE = re.compile(r"^M[A-Z0-9]{8}$")

# 1-indexed 행/열 번호 (openpyxl) — ★기본값일 뿐, 실제 위치는 `_locate()` 가 라벨로 찾는다.
#   담당자가 앞쪽에 보조 컬럼을 끼워 넣으면 통째로 밀린다(2026-08-04 실제 발생: 'F~I 채널별 일자별
#   매출목표 얼라인' 4열 삽입 → 라벨 G→K·데이터 H→L → 파서가 날짜를 못 찾고 ValueError).
_R_OPEN, _R_SELL, _R_SERIES, _R_CHANNEL, _R_STYLE, _R_PREP = 3, 4, 5, 7, 8, 10
_R_DAILY_FROM = 18
_C_LABEL = 7        # G열 = 라벨/날짜
_C_DATA_FROM = 8    # H열~ = 품번×채널

# 라벨열 판별에 쓰는 필수 라벨(이 셋이 한 열에 다 있으면 그 열이 라벨열)
_LABEL_KEYS = ("품번", "채널", "준비수량")
_ROW_LABELS = {"_R_OPEN": "판매개시일", "_R_SELL": "목표소진율", "_R_SERIES": "HERO 시리즈",
               "_R_CHANNEL": "채널", "_R_STYLE": "품번", "_R_PREP": "준비수량"}

_MD_RE = re.compile(r"(\d{1,2})\s*월\s*(\d{1,2})\s*일")


def _num(v):
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return f


def _base(style) -> str:
    return str(style).strip().split("-")[0]


def _locate(ws) -> dict:
    """라벨열·헤더행 위치를 **시트에서 직접 찾는다**(하드코딩 오프셋 금지).

    라벨열 = 위쪽 30행 안에 '품번'·'채널'·'준비수량' 이 모두 있는 열. 데이터는 그 다음 열부터.
    헤더행 = 라벨열에서 각 라벨이 실제로 있는 행. 못 찾으면 모듈 기본값으로 폴백한다.
    """
    scan_r = min(ws.max_row, 30)
    scan_c = min(ws.max_column, 40)
    c_label = None
    for c in range(1, scan_c + 1):
        labels = {str(ws.cell(r, c).value or "").strip() for r in range(1, scan_r + 1)}
        if all(k in labels for k in _LABEL_KEYS):
            c_label = c
            break
    if c_label is None:
        c_label = _C_LABEL

    rowof = {}
    for r in range(1, scan_r + 1):
        v = str(ws.cell(r, c_label).value or "").strip()
        if v and v not in rowof:
            rowof[v] = r
    loc = {k: rowof.get(lab, globals()[k]) for k, lab in _ROW_LABELS.items()}
    loc["_C_LABEL"] = c_label
    loc["_C_DATA_FROM"] = c_label + 1
    # 일별 그리드 시작 = 라벨열에서 날짜로 읽히는 첫 행(헤더 아래)
    daily_from = None
    for r in range(max(loc["_R_PREP"], _R_STYLE) + 1, min(ws.max_row, 60) + 1):
        if _cell_date(ws.cell(r, c_label).value, 2026):
            daily_from = r
            break
    loc["_R_DAILY_FROM"] = daily_from or _R_DAILY_FROM
    return loc


def _cell_date(v, year_from):
    """'07월 01일' 또는 날짜 셀 → date. 7~12월=시즌 시작연도, 1~6월=다음 해(시즌 롤오버)."""
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    m = _MD_RE.search(str(v or ""))
    if not m:
        return None
    mm, dd = int(m.group(1)), int(m.group(2))
    try:
        return datetime.date(year_from if mm >= 7 else year_from + 1, mm, dd)
    except ValueError:
        return None


def _windows(as_of, season_start):
    """as_of = 생성 기준일(오늘). 매출과 동일하게 데이터는 전일(T-1)까지."""
    if isinstance(as_of, str):
        as_of = datetime.date.fromisoformat(as_of)
    end = as_of - datetime.timedelta(days=1)
    return {
        "YTD": (season_start, end),                              # 26FW 시즌 누계(FWTD)
        "MTD": (end.replace(day=1), end),
        "WEEK": (end - datetime.timedelta(days=6), end),
        "DAY": (end, end),
    }


def _download(drive, fid):
    from googleapiclient.http import MediaIoBaseDownload
    buf = io.BytesIO()
    dl = MediaIoBaseDownload(buf, drive.files().get_media(fileId=fid, supportsAllDrives=True),
                             chunksize=4 * 1024 * 1024)
    done = False
    while not done:
        _, done = dl.next_chunk()
    buf.seek(0)
    return buf


def parse_26fw_targets(drive, as_of, fid=None, tab=TARGET_TAB) -> dict:
    """26FW 목표 시트(.xlsx) → {base: {tq, prep, sellthrough}}. 실패 시 예외(호출측에서 폴백)."""
    import openpyxl

    wb = openpyxl.load_workbook(_download(drive, fid or TARGET_FID), read_only=False, data_only=True)
    if tab not in wb.sheetnames:
        raise KeyError(f"'{tab}' 탭 없음 (탭: {wb.sheetnames})")
    ws = wb[tab]

    loc = _locate(ws)
    r_open, r_sell = loc["_R_OPEN"], loc["_R_SELL"]
    r_chan, r_style, r_prep = loc["_R_CHANNEL"], loc["_R_STYLE"], loc["_R_PREP"]
    c_label, c_data, r_daily = loc["_C_LABEL"], loc["_C_DATA_FROM"], loc["_R_DAILY_FROM"]

    # 열 메타: 데이터 열 → (base, 채널키)
    colmeta: dict[int, tuple[str, str]] = {}
    col_open: dict[int, object] = {}          # 열 → 판매개시일 원본 셀값
    for c in range(c_data, ws.max_column + 1):
        sty = str(ws.cell(r_style, c).value or "").strip()
        if not STYLE_RE.match(sty):
            continue
        chan = str(ws.cell(r_chan, c).value or "").strip().lower()
        if chan.startswith("online"):
            kch = "on"
        elif chan.startswith("offline"):
            kch = "off"
        else:
            continue
        colmeta[c] = (_base(sty), kch)
        col_open[c] = ws.cell(r_open, c).value      # 판매개시일(참고용, 누적 게이트로는 쓰지 않음)
    if not colmeta:
        raise ValueError(f"품번×채널 열을 못 찾음 — 시트 구조 변경 의심 (라벨열 {c_label}, 품번행 {r_style})")

    acc: dict[str, dict] = {}

    def slot(base):
        return acc.setdefault(base, {
            "prep_on": 0.0, "prep_off": 0.0, "sell": None,
            "tq": {p: {"on": 0.0, "off": 0.0} for p in PERIODS},
        })

    for c, (base, kch) in colmeta.items():
        sl = slot(base)
        pv = _num(ws.cell(r_prep, c).value)
        if pv:
            sl["prep_on" if kch == "on" else "prep_off"] += pv
        if sl["sell"] is None:
            sl["sell"] = _num(ws.cell(r_sell, c).value)

    # 일별 그리드 — 먼저 날짜를 모아 시즌 시작(=최초 일자)을 정한다(27SS로 갈아껴도 자동).
    rows = []
    year_from = None
    for r in range(r_daily, ws.max_row + 1):
        v = ws.cell(r, c_label).value
        if v is None:
            continue
        if year_from is None:
            # 파일에 연도가 없으면(예: '07월 01일') 시즌 시작연도를 as_of 로 추정.
            _as = datetime.date.fromisoformat(as_of) if isinstance(as_of, str) else as_of
            year_from = _as.year if _as.month >= 7 else _as.year - 1
        d = _cell_date(v, year_from)
        if d:
            rows.append((r, d))
    if not rows:
        raise ValueError(f"일별 목표 행을 못 찾음 — 시트 구조 변경 의심 "
                         f"(라벨열 {c_label}, 일별 시작행 {r_daily})")

    season_start = min(d for _, d in rows)
    windows = _windows(as_of, season_start)
    # ★판매개시일(R3)은 게이트로 쓰지 않는다 — 사용자 확정(2026-07-30): 시트 값 그대로 합산.
    #   개시일 전에도 반올림 잔여(하루 1개씩)가 있어 발매 전 히어로 달성율이 과대(라이트다운 1494%)로
    #   보일 수 있으나, 화면은 원천과 100% 일치시키고 원천 수정으로 정상화한다.
    for r, d in rows:
        inper = [p for p, (s, e) in windows.items() if s <= d <= e]
        if not inper:
            continue
        for c, (base, kch) in colmeta.items():
            v = _num(ws.cell(r, c).value)
            if not v:
                continue
            tq = acc[base]["tq"]
            for p in inper:
                tq[p][kch] += v

    out: dict[str, dict] = {}
    for base, sl in acc.items():
        on, off = sl["prep_on"], sl["prep_off"]
        out[base] = {
            "tq": {p: {"t": round(sl["tq"][p]["on"] + sl["tq"][p]["off"]),
                       "o": round(sl["tq"][p]["on"]),
                       "f": round(sl["tq"][p]["off"])} for p in PERIODS},
            "prep": {"t": round(on + off) or None, "o": round(on) or None, "f": round(off) or None},
            "sellthrough": sl["sell"],
        }
    out["_meta"] = {"season_start": season_start.isoformat(),
                    "windows": {p: [s.isoformat(), e.isoformat()] for p, (s, e) in windows.items()},
                    "styles": len(out)}
    return out


def style_prep_map(targets: dict) -> dict:
    """parse_26fw_targets 결과 → sales_rollup 의 style_prep 형식({base: {t,o,f}})."""
    return {b: {k: (t["prep"].get(k) or 0) for k in ("t", "o", "f")}
            for b, t in targets.items() if b != "_meta" and t.get("prep")}


if __name__ == "__main__":   # 단독 실행 = 파싱 + 검증
    import sys
    from pathlib import Path
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
    from soo.auth import build_services, get_credentials

    ROOT = Path(__file__).resolve().parents[2]
    drive = build_services(get_credentials(ROOT / "credentials.json", ROOT / "token.json"))["drive"]
    as_of = sys.argv[1] if len(sys.argv) > 1 else datetime.date.today().isoformat()
    t = parse_26fw_targets(drive, as_of)
    meta = t.pop("_meta")
    print(f"as_of {as_of} · 시즌시작 {meta['season_start']} · 스타일 {len(t)}")
    for p, (s, e) in meta["windows"].items():
        print(f"  {p}: {s} ~ {e}")
    for b, v in sorted(t.items(), key=lambda kv: -kv[1]["tq"]["YTD"]["t"]):
        print(f"  {b} 누계목표 {v['tq']['YTD']['t']:>7,} (on {v['tq']['YTD']['o']:,} / off {v['tq']['YTD']['f']:,})"
              f" · 준비 {v['prep']['t']:,} · 소진율목표 {v['sellthrough']}")
