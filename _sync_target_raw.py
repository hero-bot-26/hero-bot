# -*- coding: utf-8 -*-
"""목표 진실소스 통일 — xlsx `26FW HERO 일자별 목표 셋팅` → 대시보드 `히어로목표(거래량)` 탭.

배경(2026-08-10): 같은 '26FW 목표'를 두 곳이 따로 들고 있었다.
  · 앱   = Drive xlsx `26FW HERO 일자별 목표 셋팅` (담당자가 실제로 관리하는 판)
  · 시트 = 대시보드 `히어로목표(거래량)` 탭 — ★IMPORTRANGE 가 아니라 **손으로 박은 값**
같은 창(8/3~8/9)으로 대조하니 커브드팬츠 WEEK 목표가 3,629 vs 3,636 으로 갈렸다.
담당자가 xlsx 만 고치면 대시보드는 영원히 옛 값을 보여준다. → **xlsx 를 정본으로 고정**하고
이 도구가 시트를 따라오게 만든다.

★안전 규칙 (전부 실제 사고 이력에서 나온 것)
  1. **소문자 `online`/`offline` 블록만** 건드린다. 대문자 `Online`/`Offline` 블록은 2~6월
     26SS 기간 목표라 xlsx 에 없다 — 덮으면 상반기 목표가 통째로 0 이 된다.
  2. **열·행을 추가하거나 지우지 않는다.** 시트가 없는 품번×채널은 기입하지 않고 '미배치'로 보고만 한다
     (열을 끼워 넣으면 품목 탭의 SUMPRODUCT `$C$2:$GV$2` 범위가 밀린다).
  3. 값이 같은 셀은 건드리지 않는다 → 재실행하면 **0건**(멱등).
  4. `--dry` 가 기본. `--apply` 시 직전값을 JSON 으로 백업하고, 쓴 뒤 되읽어 검증한다.

    python _sync_target_raw.py            # 대조만
    python _sync_target_raw.py --apply    # 기입
"""
from __future__ import annotations

import argparse
import datetime
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

from soo.auth import build_services, get_credentials                      # noqa: E402
from soo.hero_ops import target_26fw as T                                 # noqa: E402

DASH_SID = "1-A04_TwKZJNPkFg27USkKAScZRu6CAhbgVeXk9c09nA"   # 26FW 히어로 실적 대시보드
RAW_TAB = "히어로목표(거래량)"
R_CHANNEL, R_STYLE, R_DAILY_FROM, R_DAILY_TO = 2, 3, 14, 378
C_DATE = 2                       # B열 = 일자(시리얼)
MAX_COL_LETTER = "GV"            # 품목 탭 SUMPRODUCT 가 훑는 마지막 열


def col_name(idx0: int) -> str:
    s, i = "", idx0 + 1
    while i:
        i, r = divmod(i - 1, 26)
        s = chr(65 + r) + s
    return s


def serial(d: datetime.date) -> int:
    return (d - datetime.date(1899, 12, 30)).days


def xlsx_daily(drive) -> tuple[dict, tuple]:
    """xlsx → {(base, 'online'|'offline'): {date: qty}} + 날짜 범위."""
    import openpyxl
    wb = openpyxl.load_workbook(T._download(drive, T.TARGET_FID), read_only=False, data_only=True)
    ws = wb[T.TARGET_TAB]
    loc = T._locate(ws)
    r_chan, r_style = loc["_R_CHANNEL"], loc["_R_STYLE"]
    c_label, c_data, r_daily = loc["_C_LABEL"], loc["_C_DATA_FROM"], loc["_R_DAILY_FROM"]

    colmeta = {}
    for c in range(c_data, ws.max_column + 1):
        sty = str(ws.cell(r_style, c).value or "").strip()
        if not T.STYLE_RE.match(sty):
            continue
        chan = str(ws.cell(r_chan, c).value or "").strip().lower()
        if chan.startswith("online"):
            colmeta[c] = (T._base(sty), "online")
        elif chan.startswith("offline"):
            colmeta[c] = (T._base(sty), "offline")
    if not colmeta:
        raise ValueError("xlsx 품번×채널 열을 못 찾음 — 레이아웃 변경 의심")

    out, dates = {}, []
    year_from = None
    for r in range(r_daily, ws.max_row + 1):
        v = ws.cell(r, c_label).value
        if v is None:
            continue
        if year_from is None:
            year_from = 2026
        d = T._cell_date(v, year_from)
        if not d:
            continue
        dates.append(d)
        for c, key in colmeta.items():
            out.setdefault(key, {})[d] = T._num(ws.cell(r, c).value) or 0
    return out, (min(dates), max(dates))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()
    sys.stdout.reconfigure(encoding="utf-8")

    sv = build_services(get_credentials(ROOT / "credentials.json", ROOT / "token.json"))
    sheets, drive = sv["sheets"], sv["drive"]

    src, (d_from, d_to) = xlsx_daily(drive)
    styles = sorted({b for b, _ in src})
    print(f"[xlsx] 품번 {len(styles)}개 × 채널 / 일자 {d_from} ~ {d_to}")

    grid = sheets.spreadsheets().values().get(
        spreadsheetId=DASH_SID, range=f"'{RAW_TAB}'!A1:{MAX_COL_LETTER}{R_DAILY_TO}",
        valueRenderOption="UNFORMATTED_VALUE").execute().get("values", [])

    def row(n):
        r = grid[n - 1] if len(grid) >= n else []
        return r + [""] * (250 - len(r))

    chan_row, style_row = row(R_CHANNEL), row(R_STYLE)
    # ★소문자 블록만. 대문자는 26SS 상반기라 건드리면 안 된다.
    dest = {}
    for ci, sty in enumerate(style_row):
        s = str(sty).strip()
        if not T.STYLE_RE.match(s):
            continue
        ch = str(chan_row[ci]).strip()
        if ch in ("online", "offline"):
            dest[(T._base(s), ch)] = ci

    row_of_date = {}
    for rn in range(R_DAILY_FROM, R_DAILY_TO + 1):
        v = row(rn)[C_DATE - 1]
        if isinstance(v, (int, float)) and v:
            row_of_date[datetime.date(1899, 12, 30) + datetime.timedelta(days=int(v))] = rn

    missing = sorted(k for k in src if k not in dest)
    if missing:
        print(f"[미배치] 시트에 열이 없는 품번×채널 {len(missing)}건 — 기입하지 않음(열 삽입 금지):")
        for b, c in missing[:12]:
            print(f"    {b} {c}")

    updates, diffs, same = [], [], 0
    for key, daily in src.items():
        ci = dest.get(key)
        if ci is None:
            continue
        for d, want in daily.items():
            rn = row_of_date.get(d)
            if rn is None:
                continue
            cur = row(rn)[ci]
            cur = float(cur) if isinstance(cur, (int, float)) else 0.0
            if abs(cur - float(want)) < 1e-9:
                same += 1
                continue
            a1 = f"'{RAW_TAB}'!{col_name(ci)}{rn}"
            updates.append({"range": a1, "values": [[want]]})
            diffs.append((key[0], key[1], d.isoformat(), cur, want))

    print(f"\n[대조] 일치 {same:,}셀 / 불일치 {len(diffs):,}셀")
    if diffs:
        by_style = {}
        for b, c, d, cur, want in diffs:
            k = by_style.setdefault(b, {"n": 0, "cur": 0.0, "want": 0.0, "first": d, "last": d})
            k["n"] += 1; k["cur"] += cur; k["want"] += want
            k["first"] = min(k["first"], d); k["last"] = max(k["last"], d)
        print(f"{'품번':12} {'셀':>6} {'현재합':>10} {'xlsx합':>10}   기간")
        for b, v in sorted(by_style.items()):
            print(f"  {b:12} {v['n']:>6} {v['cur']:>10,.0f} {v['want']:>10,.0f}   {v['first']}~{v['last']}")

    if not updates:
        print("\n[OK] 시트가 이미 xlsx 와 같습니다 (멱등).")
        return
    if not args.apply:
        print("\n드라이런입니다. 실제로 기입하려면 --apply 를 붙이세요.")
        return

    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    bak = ROOT / f"_target_raw_backup_{stamp}.json"
    bak.write_text(json.dumps({"cells": [{"range": u["range"], "prev": d[3]}
                                         for u, d in zip(updates, diffs)]},
                              ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"\n[백업] {bak}")

    # ★사내망(VDI)이 큰 배치에서 소켓 타임아웃을 자주 낸다(실제로 밟음) — 잘게 쪼개고 직접 재시도한다.
    #   멱등이라 중간에 끊겨도 다시 돌리면 남은 것만 쓴다.
    import time
    CHUNK = 1200
    for i in range(0, len(updates), CHUNK):
        part = updates[i:i + CHUNK]
        for attempt in range(5):
            try:
                sheets.spreadsheets().values().batchUpdate(spreadsheetId=DASH_SID, body={
                    "valueInputOption": "RAW", "data": part}).execute()
                break
            except Exception as e:
                if attempt == 4:
                    raise
                wait = 3 * (attempt + 1)
                print(f"    재시도 {attempt + 1}/4 ({type(e).__name__}) — {wait}s 후")
                time.sleep(wait)
        print(f"  기입 {min(i + CHUNK, len(updates)):,}/{len(updates):,}")
    print("[적용] 완료 — 되읽어 검증합니다")

    grid2 = sheets.spreadsheets().values().get(
        spreadsheetId=DASH_SID, range=f"'{RAW_TAB}'!A1:{MAX_COL_LETTER}{R_DAILY_TO}",
        valueRenderOption="UNFORMATTED_VALUE").execute().get("values", [])

    def row2(n):
        r = grid2[n - 1] if len(grid2) >= n else []
        return r + [""] * (250 - len(r))

    left = 0
    for key, daily in src.items():
        ci = dest.get(key)
        if ci is None:
            continue
        for d, want in daily.items():
            rn = row_of_date.get(d)
            if rn is None:
                continue
            cur = row2(rn)[ci]
            cur = float(cur) if isinstance(cur, (int, float)) else 0.0
            if abs(cur - float(want)) >= 1e-9:
                left += 1
    print("[검증] 되읽기 불일치:", left, "건")


if __name__ == "__main__":
    main()
